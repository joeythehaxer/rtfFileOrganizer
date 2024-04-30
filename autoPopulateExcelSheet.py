import os
import openpyxl
from striprtf.striprtf import rtf_to_text
import re
from langdetect import detect

# Constants
FOLDER_PATH = r"C ACTIONS TO ADD"
EXCEL_FILE_PATH = r"New Microsoft Excel Worksheet.xlsx"
SHEET_NAME = "Sheet1"


def remove_non_english(text):

    # Tokenize the text into words
    words = re.findall(r'\b\w+\b', text)

    # Filter out non-English words
    english_words = [word for word in words if any(c.isalpha() for c in word) and word != "FDI"]

    # Join the English words into a sentence
    english_sentence = ' '.join(english_words)

    return english_sentence


def summarize(text):
    # Remove non-English text
    english_text = remove_non_english(text)

    # Return the entire English text
    return english_text


def rephrase(text):
    # Split the text into sentences
    sentences = re.split(r'(?<!\w\.\w.)(?<![A-Z][a-z]\.)(?<=\.|\?)\s', text)

    # Rephrase each sentence
    rephrased_sentences = []
    for sentence in sentences:
        # Split the sentence into words
        words = sentence.split()

        # Remove non-English words
        english_words = [word for word in words]

        # Join the English words into a sentence
        rephrased_sentence = ' '.join(english_words)

        rephrased_sentences.append(rephrased_sentence)

    # Join the rephrased sentences into a single string
    rephrased_text = ' '.join(rephrased_sentences)

    return rephrased_text


def load_excel_workbook(excel_file_path):
    return openpyxl.load_workbook(excel_file_path)


def find_last_empty_row(ws):
    return ws.max_row + 1


def extract_info_from_rtf(file_path):
    extracted_info = []
    with open(file_path, 'r') as f:
        rtf_content = f.read()
        plain_text = rtf_to_text(rtf_content)
        lines = plain_text.split("|")
        clientraw = lines[0].split(">")[1]
        client = ""
        for char in clientraw:
            if char.isalnum() or char == " ":
                client += char
        clientDerivative = lines[50]
        jobString = "3197"
        poNumberString = lines[2]
        addressString = lines[23]
        postcodePattern = re.compile(r'\b[A-Z]{1,2}\d{1,2}\s\d[A-Z]{2}\b')
        riskString = ""
        try:
            postcodeString = postcodePattern.findall(addressString)[0]
        except IndexError or UnboundLocalError:
            postcodeString = "None"
        telString = "NOT SUPPLIED"
        if lines[16] != "":
            telString = lines[16]
        dateRaisedString = lines[38]
        targetDateString = lines[46]
        tenderNumberPattern = re.compile(r'\b\d{5,7}\b')
        tenderNumberList = tenderNumberPattern.findall(lines[61])

        # print(tenderNumberList)
        try:
            tenderNumberString = tenderNumberList[0]
        except IndexError:
            tenderNumberString = "None"
        rawActionDescriptionString = lines[61]
        englishActionString = summarize(rawActionDescriptionString)


        if len(tenderNumberList) <= 1:
            extracted_info.append(clientDerivative)
            extracted_info.append(jobString)
            extracted_info.append(poNumberString)
            extracted_info.append(tenderNumberString)
            extracted_info.append(dateRaisedString)
            extracted_info.append(targetDateString)
            extracted_info.append(telString)
            extracted_info.append(addressString)
            extracted_info.append(postcodeString)
            extracted_info.append(riskString)
            extracted_info.append(englishActionString)
            return extracted_info
        else:
            extracted_info.append(clientDerivative)
            extracted_info.append(jobString)
            extracted_info.append(poNumberString)
            extracted_info.append(tenderNumberList)
            extracted_info.append(dateRaisedString)
            extracted_info.append(targetDateString)
            extracted_info.append(telString)
            extracted_info.append(addressString)
            extracted_info.append(postcodeString)
            extracted_info.append(riskString)
            extracted_info.append(englishActionString)
            return extracted_info


def paste_info_to_excel(ws, extracted_info, last_row):
    reProcessedExtractedInfo = []
    for i in extracted_info:
        reProcessedExtractedInfo.append(i)

    if str(type(extracted_info[3])) != "<class \'list\'>":
        for i, cell_text in enumerate(extracted_info):
            ws.cell(row=last_row, column=i + 2).value = cell_text
        return last_row + 1
    else:
        v = 0
        for i in extracted_info[3]:
            reProcessedExtractedInfo[3] = extracted_info[3][v]
            # print(i)
            for x, cell_text in enumerate(reProcessedExtractedInfo):
                ws.cell(row=last_row, column=x + 2).value = cell_text
            v += 1
            last_row += 1
        return last_row


def save_excel_workbook(wb, excel_file_path):
    wb.save(excel_file_path)


def import_info_from_rtf_folder(folder_path, excel_file_path, sheet_name):
    # Load Excel workbook
    wb = load_excel_workbook(excel_file_path)
    ws = wb[sheet_name]

    # Loop through each RTF file in the folder
    for filename in os.listdir(folder_path):
        if filename.endswith(".rtf"):
            # Construct the full file path
            # print(filename)
            file_path = os.path.join(folder_path, filename)

            # Extract information from RTF file
            extracted_info = extract_info_from_rtf(file_path)

            # Find the last empty row in Excel sheet
            last_row = find_last_empty_row(ws)

            # Paste the extracted information into a new row in Excel
            paste_info_to_excel(ws, extracted_info, last_row)

    # Save changes to Excel workbook
    save_excel_workbook(wb, excel_file_path)


# Usage
import_info_from_rtf_folder(FOLDER_PATH, EXCEL_FILE_PATH, SHEET_NAME)
